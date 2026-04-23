# accounts/management/commands/export_customers_xlsx.py
from pathlib import Path
from datetime import datetime

from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from django.utils.timezone import localtime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, numbers

from accounts.models import UserPackage, Payment


def _style_header(row):
    """Bold, center header cells."""
    for cell in row:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws):
    """Autosize columns based on contents (cap width at 50)."""
    for col in ws.columns:
        max_len = 0
        first = next(iter(col), None)
        if not first:
            continue
        col_letter = first.column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)


User = get_user_model()


class Command(BaseCommand):
    help = "Export customers, user packages, and payments to a single Excel workbook."

    def add_arguments(self, parser):
        parser.add_argument(
            "--all-users",
            action="store_true",
            help="Include all users (not only those with packages).",
        )

    def handle(self, *args, **opts):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        outdir = Path("exports")
        outdir.mkdir(parents=True, exist_ok=True)
        xlsx_path = outdir / f"customers_{ts}.xlsx"

        # --- Choose the user scope
        if opts["all_users"]:
            users_qs = User.objects.all().order_by("id")
        else:
            users_qs = (
                User.objects.filter(userpackage__isnull=False)
                .distinct()
                .order_by("id")
            )

        # Prepare ids to filter related tables when not exporting all users
        user_ids = list(users_qs.values_list("id", flat=True))

        packages_qs = (
            UserPackage.objects.select_related("user", "package").order_by("id")
        )
        payments_qs = (
            Payment.objects.select_related("user", "user_package", "user_package__package")
            .order_by("id")
        )

        if not opts["all_users"]:
            packages_qs = packages_qs.filter(user_id__in=user_ids)
            payments_qs = payments_qs.filter(user_id__in=user_ids)

        # --- Build workbook
        wb = Workbook()

        # Sheet 1: Users
        ws_users = wb.active
        ws_users.title = "Users"
        ws_users.append(
            [
                "user_id",
                "username",
                "email",
                "first_name",
                "last_name",
                "is_active",
                "date_joined",
            ]
        )
        _style_header(ws_users[1])
        ws_users.freeze_panes = "A2"

        for u in users_qs:
            dt = getattr(u, "date_joined", None)
            dt_local = localtime(dt) if dt else None
            ws_users.append(
                [
                    u.id,
                    getattr(u, "username", "") or "",
                    getattr(u, "email", "") or "",
                    getattr(u, "first_name", "") or "",
                    getattr(u, "last_name", "") or "",
                    bool(u.is_active),
                    dt_local.replace(tzinfo=None) if dt_local else "",
                ]
            )

        # Apply formats for date column (G) after data
        for r in range(2, ws_users.max_row + 1):
            ws_users[f"G{r}"].number_format = "yyyy-mm-dd hh:mm"

        ws_users.auto_filter.ref = ws_users.dimensions
        _autofit(ws_users)

        # Sheet 2: UserPackages
        ws_up = wb.create_sheet("UserPackages")
        ws_up.append(
            [
                "userpackage_id",
                "user_id",
                "username",
                "package_slug",
                "package_name",
                "package_type",
                "status",
                "step",
                "price_usd",
                "paid_usd",
                "due_usd",
                "created_at",
            ]
        )
        _style_header(ws_up[1])
        ws_up.freeze_panes = "A2"

        for up in packages_qs:
            created_local = localtime(up.created_at) if up.created_at else None
            # Convert cents -> dollars
            price_usd = (up.package.price_cents or 0) / 100.0
            paid_usd = (up.paid_cents or 0) / 100.0
            due_usd = (up.due_cents or 0) / 100.0

            ws_up.append(
                [
                    up.id,
                    up.user_id,
                    getattr(up.user, "username", "") or "",
                    up.package.slug,
                    up.package.name,
                    up.package.type,
                    up.status,
                    up.step,
                    price_usd,
                    paid_usd,
                    due_usd,
                    created_local.replace(tzinfo=None) if created_local else "",
                ]
            )

            last = ws_up.max_row
            ws_up[f"I{last}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            ws_up[f"J{last}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            ws_up[f"K{last}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            ws_up[f"L{last}"].number_format = "yyyy-mm-dd hh:mm"

        ws_up.auto_filter.ref = ws_up.dimensions
        _autofit(ws_up)

        # Sheet 3: Payments
        ws_pay = wb.create_sheet("Payments")
        ws_pay.append(
            [
                "payment_id",
                "user_id",
                "username",
                "userpackage_id",
                "package_slug",
                "package_name",
                "amount_usd",
                "status",
                "created_at",
            ]
        )
        _style_header(ws_pay[1])
        ws_pay.freeze_panes = "A2"

        for p in payments_qs:
            created_local = localtime(p.created_at) if p.created_at else None
            amount_usd = (p.amount_cents or 0) / 100.0

            ws_pay.append(
                [
                    p.id,
                    p.user_id,
                    getattr(p.user, "username", "") or "",
                    p.user_package_id,
                    p.user_package.package.slug
                    if p.user_package and p.user_package.package
                    else "",
                    p.user_package.package.name
                    if p.user_package and p.user_package.package
                    else "",
                    amount_usd,
                    p.status,
                    created_local.replace(tzinfo=None) if created_local else "",
                ]
            )

            last = ws_pay.max_row
            ws_pay[f"G{last}"].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            ws_pay[f"I{last}"].number_format = "yyyy-mm-dd hh:mm"

        ws_pay.auto_filter.ref = ws_pay.dimensions
        _autofit(ws_pay)

        # Save once at the end
        wb.save(xlsx_path)
        self.stdout.write(self.style.SUCCESS(f"Excel export complete: {xlsx_path}"))
