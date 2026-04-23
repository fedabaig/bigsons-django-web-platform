# accounts/management/commands/export_customers_xlsx.py
from pathlib import Path
from datetime import datetime
from django.core.management.base import BaseCommand
from django.contrib.auth import get_user_model
from django.utils.timezone import localtime
from openpyxl import Workbook

from accounts.models import UserPackage, Payment

User = get_user_model()

class Command(BaseCommand):
    help = "Export customers, user packages, and payments to a single Excel workbook."

    def add_arguments(self, parser):
        parser.add_argument("--all-users", action="store_true", help="Include all users.")

    def handle(self, *args, **opts):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        outdir = Path("exports")
        outdir.mkdir(parents=True, exist_ok=True)
        xlsx_path = outdir / f"customers_{ts}.xlsx"

        if opts["all_users"]:
            users_qs = User.objects.all().order_by("id")
        else:
            users_qs = (User.objects
                        .filter(userpackage__isnull=False)
                        .distinct()
                        .order_by("id"))

        packages_qs = (UserPackage.objects
                       .select_related("user", "package")
                       .order_by("id"))
        payments_qs = (Payment.objects
                       .select_related("user", "user_package", "user_package__package")
                       .order_by("id"))

        wb = Workbook()

        # Sheet 1: Users
        ws_users = wb.active
        ws_users.title = "Users"
        ws_users.append(["user_id", "username", "email", "first_name", "last_name", "is_active", "date_joined"])
        for u in users_qs:
            ws_users.append([
                u.id,
                getattr(u, "username", ""),
                getattr(u, "email", ""),
                getattr(u, "first_name", ""),
                getattr(u, "last_name", ""),
                u.is_active,
                localtime(getattr(u, "date_joined", None)).strftime("%Y-%m-%d %H:%M:%S") if getattr(u, "date_joined", None) else "",
            ])

        # Sheet 2: UserPackages
        ws_up = wb.create_sheet("UserPackages")
        ws_up.append([
            "userpackage_id", "user_id", "username",
            "package_slug", "package_name", "package_type",
            "status", "step", "price_cents", "paid_cents", "due_cents", "created_at",
        ])
        for up in packages_qs:
            ws_up.append([
                up.id,
                up.user_id,
                getattr(up.user, "username", ""),
                up.package.slug,
                up.package.name,
                up.package.type,
                up.status,
                up.step,
                up.package.price_cents,
                up.paid_cents,
                up.due_cents,
                localtime(up.created_at).strftime("%Y-%m-%d %H:%M:%S") if up.created_at else "",
            ])

        # Sheet 3: Payments
        ws_pay = wb.create_sheet("Payments")
        ws_pay.append([
            "payment_id", "user_id", "username",
            "userpackage_id", "package_slug", "package_name",
            "amount_cents", "status", "created_at",
        ])
        for p in payments_qs:
            ws_pay.append([
                p.id,
                p.user_id,
                getattr(p.user, "username", ""),
                p.user_package_id,
                p.user_package.package.slug if p.user_package and p.user_package.package else "",
                p.user_package.package.name if p.user_package and p.user_package.package else "",
                p.amount_cents,
                p.status,
                localtime(p.created_at).strftime("%Y-%m-%d %H:%M:%S") if p.created_at else "",
            ])

        wb.save(xlsx_path)
        self.stdout.write(self.style.SUCCESS(f"Excel export complete: {xlsx_path}"))
